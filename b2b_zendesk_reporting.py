"""
B2C Zendesk Reporting Module
─────────────────────────────
Categorizes ALL Zendesk tickets by Issue text using the same
MEDICAL_DEVICE_CATEGORIES as the Return Categorizer, then reports
only quality-relevant tickets aggregated by Parent SKU (first 7 chars).

Categorization modes:
  • Keyword matching (free, instant)
  • AI for unclear tickets + random audit sample
  • Full AI mode (optional)
"""

import streamlit as st
import pandas as pd
import numpy as np
import re
from datetime import datetime, timedelta
from io import BytesIO
from typing import Optional, Dict, List, Tuple
import logging
import random

logger = logging.getLogger(__name__)

# Optional AI import — gracefully absent if not available
try:
    from enhanced_ai_analysis import EnhancedAIAnalyzer, AIProvider
    _AI_IMPORT_OK = True
except ImportError:
    _AI_IMPORT_OK = False

# ─── Constants ───────────────────────────────────────────────────────────────

PARENT_SKU_LENGTH = 7

EXCLUDED_SKUS = {"x", ".", "X", "XX", "XXX", "No SKU", "", " "}

NOISE_ISSUES = {
    "x", "X", "XX", "XXX", ".", " ", "",
    "Amazon calls.", "AMAZON CALL SIDE TASK", "Amazon calls",
    "Amazon call side tasks", "Amazon call side task",
    "Amazon call", "Amazon Call", "Amazon call.",
    "order status", "Order status", "Order status.",
    "delivery status", "delivery", "delivery SLA", "delivery sla",
    "Delivery update.", "Delivery address update.",
    "Cancelation", "discount code", "call dropped",
    "Missed chat.", "Miscellaneous.", "Miscellaneous",
    "Missed chat due back to back calls. No email to provide an answer.",
    "Call returned in 623535", "Callback not answered",
}

# ─── MEDICAL DEVICE CATEGORIES ────────────────────────────────────────────────

QUALITY_CATEGORIES = [
    'Size: Too Small',
    'Size: Too Large',
    "Size: Doesn't Fit / Wrong Dimensions",
    'Comfort: Causes Pain or Pressure',
    'Comfort: Too Hard / Rigid',
    'Comfort: Too Soft / Lacks Support',
    'Comfort: Skin Irritation or Allergic Reaction',
    'Defect: Broken / Structural Failure',
    'Defect: Malfunctions / Stops Working',
    'Defect: Cosmetic Damage',
    'Defect: Poor Material Quality',
    'Wrong Product / Not as Described',
    'Missing or Incomplete Components',
    "Performance: Ineffective / Doesn't Help",
    'Equipment Compatibility Issue',
    'Stability: Shifts / Unstable / Falls',
    'Assembly / Usage Difficulty',
    'Medical / Safety Concern',
]

NON_QUALITY_CATEGORIES = [
    'Customer: Changed Mind / No Longer Needed',
    'Customer: Ordered Wrong Size or Item',
    'Fulfillment: Damaged in Shipping',
    'Fulfillment: Wrong Item Sent',
    'Fulfillment: Delivery Issue',
    'General Inquiry / Not a Quality Issue',
    'Other / Miscellaneous',
]

ALL_CATEGORIES = QUALITY_CATEGORIES + NON_QUALITY_CATEGORIES

INJURY_KEYWORDS = [
    'injur', 'hospital', 'emergency', 'fell', 'fall', 'tip over',
    'tipped', 'dangerous', 'unsafe', 'hazard', 'hurt', 'accident',
    'cut', 'bleed', 'burn', 'shock', 'pinch', 'trap',
]

# ─── KEYWORD RULES ────────────────────────────────────────────────────────────

KEYWORD_RULES: List[Tuple[str, List[str]]] = [
    ('Medical / Safety Concern', [
        'injury', 'injured', 'hospital', 'emergency', 'dangerous', 'unsafe',
        'hazard', 'fell off', 'fall', 'fell', 'tipping', 'tipped over',
        'cut my leg', 'cut myself', 'hurt', 'accident',
    ]),
    ('Defect: Broken / Structural Failure', [
        'broke', 'broken', 'snapped', 'cracked', 'crack', 'shattered',
        'fell apart', 'ripped', 'rip', 'torn', 'tore', 'split',
        'bent', 'buckle', 'weld', 'disconnected from the frame',
        'collapsed', 'structural', 'frame broke', 'clamp broke',
        'handle broke', 'leg broke', 'arm broke', 'pin broke',
        'connector broke', 'broken clamp', 'broken connector',
        'broken handle', 'broken part', 'broken piece', 'broken battery',
        'broke off', 'snapped off',
    ]),
    ('Defect: Malfunctions / Stops Working', [
        'malfunction', 'stops working', 'stopped working', 'not working',
        "won't turn on", 'wont turn on', "won't work", 'wont work',
        "doesn't work", 'doesnt work', 'not turning on', "won't operate",
        'quit working', 'dead', 'beeping', 'beeps', '2 beeps',
        'not charging', 'not taking a charge', "won't charge",
        'motor', 'control box defective', 'defective',
        'not running', "won't run", 'stops when riding',
        'stops dead', 'power issue', 'flashing', 'alarm sound',
        'not getting cold', "isn't working", 'pump not working',
        'pump quit', 'pump dead', 'pump making loud',
        'not holding air', 'losing power', 'not turning',
        'joystick', 'charger', 'not operate', 'scooter issue',
    ]),
    ('Stability: Shifts / Unstable / Falls', [
        'wobbly', 'wobble', 'wobbling', 'unstable', 'shifts',
        'slides', 'tipping forward', 'not stable', 'not secure',
        'tilted', 'not locking', "won't lock", 'wont lock',
        "doesn't lock", 'not stay', "won't stay",
        'seat not locking', 'lock in place',
    ]),
    ('Defect: Poor Material Quality', [
        'cheap', 'poor quality', 'low quality', 'thin',
        'flimsy', 'wear out', 'worn out', 'wears out',
        'peeling', 'discolor', 'rusting', 'rust',
        'velcro stopped', 'fabric', 'material',
        'odor', 'smell', 'stain',
    ]),
    ('Missing or Incomplete Components', [
        'missing part', 'missing piece', 'missing hardware',
        'missing bolt', 'missing screw', 'missing nut',
        'missing washer', 'missing component', 'missing gel',
        'missing horn', 'missing tip', 'missing the',
        'missing a', 'missing item', 'parts missing',
        'incomplete', 'no instructions', 'without instructions',
        'empty package', 'package was empty', 'not included',
        "didn't include", 'only received 1', 'only received one',
        'missing push pin', 'missing handscrew', 'missing leg',
        'missing axle', 'missing harware', 'missing hardware kit',
    ]),
    ('Comfort: Causes Pain or Pressure', [
        'pain', 'painful', 'pressure', 'sore', 'bruise', 'bruising',
        'digs in', 'rubs', 'rubbing', 'uncomfortable', 'discomfort',
        'causes pain', 'hurts', 'numbed',
    ]),
    ('Comfort: Too Hard / Rigid', [
        'too hard', 'too stiff', 'too rigid', 'really stiff',
        'extremely hard', 'too firm',
    ]),
    ('Comfort: Too Soft / Lacks Support', [
        'too soft', 'no support', 'lacks support', 'deflated',
        'collapsed under', 'not enough support', 'stayed deflated',
        'sinking', 'not expand',
    ]),
    ('Comfort: Skin Irritation or Allergic Reaction', [
        'rash', 'irritation', 'allergic', 'skin reaction', 'itching', 'itch',
    ]),
    ('Size: Too Small', [
        'too small', 'too short', 'too narrow', 'too tight',
    ]),
    ('Size: Too Large', [
        'too large', 'too big', 'too wide', 'too long', 'too bulky',
        'too heavy', 'oversized',
    ]),
    ("Size: Doesn't Fit / Wrong Dimensions", [
        "doesn't fit", "doesnt fit", "does not fit", "not fit",
        'not fitting', "won't fit", "didn't fit", 'wrong size',
        'wrong dimension', 'bad fit', 'good fit',
    ]),
    ("Performance: Ineffective / Doesn't Help", [
        'ineffective', "doesn't help", "doesn't do anything",
        "doesn't work for", 'not effective', 'useless',
        "didn't help", 'not able to use it',
    ]),
    ('Equipment Compatibility Issue', [
        'not compatible', 'incompatible', "doesn't attach",
        'does not fit my walker', 'does not fit my wheelchair',
        "doesn't connect", 'not pairing',
    ]),
    ('Assembly / Usage Difficulty', [
        'hard to assemble', 'difficult to assemble', 'assembly issue',
        'parts not fitting', 'unable to assemble', 'confusing instruction',
        'how to', 'help assembl', 'help putting',
        'not able to open', 'problem assembling',
    ]),
    ('Defect: Cosmetic Damage', [
        'scratch', 'scratched', 'dent', 'dented', 'cosmetic',
        'paint', 'hubcap', 'dust on',
    ]),
    ('Wrong Product / Not as Described', [
        'wrong product', 'wrong item', 'not as described',
        'incorrect product', 'incorrect item', 'received the incorrect',
        'wrong package', "didn't order", 'not what I ordered',
        'different product', 'different color',
    ]),
    ('Fulfillment: Damaged in Shipping', [
        'damaged in shipping', 'damaged during shipping',
        'box arrived damaged', 'box was damaged', 'arrived damaged',
        'shipping damage', 'damaged box', 'scuffed',
        'package was damaged', 'package arrived damaged',
    ]),
    ('Fulfillment: Wrong Item Sent', [
        'sent wrong', 'shipped wrong', 'wrong item sent',
        'received wrong', 'wanted the beige and received',
    ]),
    ('Fulfillment: Delivery Issue', [
        'not delivered', 'never arrived', 'never received',
        'not received', 'lost package', 'lost order',
        'returned to sender', 'order not received',
        'has not been shipped', "hasn't been shipped",
        'item not delivered',
    ]),
    ('Customer: Changed Mind / No Longer Needed', [
        'changed mind', 'no longer need', 'no longer needed',
        'not needed', "doesn't need", 'does not need',
        'decided not to', 'not a good fit for',
        'doctor recommended not', 'doctor advised',
        'hospice', 'will not use', "won't use",
        'return product for a better price',
        'return the item', 'wants to return',
        'looking to return', 'return request',
        'return/refund request', 'return/refund operation',
        'return authorization', 'return process',
        'return and refund', 'return request not needed',
        'not meet', "didn't meet", 'expectations',
    ]),
    ('Customer: Ordered Wrong Size or Item', [
        'ordered wrong', 'purchased incorrect', 'ordered the wrong',
        'he purchased the incorrect', 'bought wrong',
    ]),
]

BRAKE_KEYWORDS = ['brake', 'brakes', 'braking']

PART_REQUEST_QUALITY_KEYWORDS = [
    'replacement part', 'needs part', 'need part', 'needs the',
    'needs a new', 'replacement wheel', 'replacement tire',
    'replacement battery', 'needs replacement',
    'part request', 'parts request',
]


# ─── Keyword Categorizer ──────────────────────────────────────────────────────

def categorize_by_keywords(issue_text: str) -> Tuple[str, float]:
    if not issue_text or not isinstance(issue_text, str):
        return 'General Inquiry / Not a Quality Issue', 0.0

    text = issue_text.lower().strip()

    if text in {s.lower() for s in NOISE_ISSUES} or len(text) < 3:
        return 'General Inquiry / Not a Quality Issue', 1.0

    if any(kw in text for kw in BRAKE_KEYWORDS):
        if any(kw in text for kw in ['not working', "doesn't", "won't", "wont",
                                      'not lock', 'not engage', 'stiff', 'hard to',
                                      'broke', 'broken', 'bent', 'issue', 'problem',
                                      'defective', 'not move', 'floppy']):
            return 'Defect: Malfunctions / Stops Working', 0.9
        return 'Defect: Malfunctions / Stops Working', 0.7

    if any(kw in text for kw in PART_REQUEST_QUALITY_KEYWORDS):
        if any(kw in text for kw in ['broke', 'broken', 'worn', 'fell off',
                                      'fell apart', 'cracked', 'snapped']):
            return 'Defect: Broken / Structural Failure', 0.8
        return 'Defect: Malfunctions / Stops Working', 0.6

    for category, keywords in KEYWORD_RULES:
        for kw in keywords:
            if kw in text:
                return category, 0.85

    # ── Broader "missing" catch (after specific missing keywords above) ──
    if 'missing' in text:
        return 'Missing or Incomplete Components', 0.7

    if any(kw in text for kw in ['only came with', 'only received',
                                  'supposed to come with', 'did not receive']):
        return 'Missing or Incomplete Components', 0.7

    if any(kw in text for kw in ['leak', 'leaking', 'leaks']):
        return 'Defect: Malfunctions / Stops Working', 0.7

    if any(kw in text for kw in ['noise', 'noisy', 'loud', 'grinding',
                                  'squeaking', 'clicking', 'squeaks']):
        return 'Defect: Malfunctions / Stops Working', 0.7

    if any(kw in text for kw in ['seized', 'stuck', 'jammed', 'cannot be loosened',
                                  'cannot adjust']):
        return 'Defect: Malfunctions / Stops Working', 0.7

    if any(kw in text for kw in ['non working', 'non-working', 'does not work',
                                  'do not work', 'not work at all']):
        return 'Defect: Malfunctions / Stops Working', 0.8

    if 'too tall' in text:
        return 'Size: Too Large', 0.8

    # "wrong color" / "wrong one" / "sent the wrong" → Wrong Product
    if any(kw in text for kw in ['wrong color', 'wrong colour', 'wrong one',
                                  'sent the wrong', 'received the wrong',
                                  'was sent the wrong']):
        return 'Wrong Product / Not as Described', 0.8

    # Battery with problem context
    if 'batter' in text:
        if any(kw in text for kw in ['not holding', 'not working', 'non working',
                                      'dead', 'issue', 'defective', 'problem',
                                      'not charging', "won't charge", 'not taking',
                                      'pushed inside', 'port pushed']):
            return 'Defect: Malfunctions / Stops Working', 0.7

    if 'fold' in text:
        if any(kw in text for kw in ['not fold', "doesn't fold", "won't fold",
                                      'issue', 'problem', 'not work', 'mechanism',
                                      'incorrectly', 'correctly']):
            return 'Defect: Malfunctions / Stops Working', 0.7

    if any(kw in text for kw in ['issue with', 'issues with', 'having issue',
                                  'having issues', 'product issue']):
        if not any(kw in text for kw in ['delivery', 'payment', 'order', 'fixed']):
            return 'Defect: Malfunctions / Stops Working', 0.5

    # "pulls to the right/left" / alignment issues
    if any(kw in text for kw in ['pulls to the', 'not pumping', 'not inflat',
                                  'not sturdy', 'rotate too much',
                                  'not engaging', 'not locking']):
        return 'Defect: Malfunctions / Stops Working', 0.7

    # "damaged when arrived" / "arrived damaged" (broader shipping damage)
    if any(kw in text for kw in ['damaged when', 'arrived damaged', 'damaged upon',
                                  'damaged on arrival']):
        return 'Fulfillment: Damaged in Shipping', 0.7

    # "breaks" misspelling of "brakes"
    if 'breaks' in text and any(kw in text for kw in ['not', "won't", "don't",
                                                       'engage', 'lock', 'work']):
        return 'Defect: Malfunctions / Stops Working', 0.7

    # "drilled on the wrong side" / manufacturing defect language
    if any(kw in text for kw in ['drilled', 'manufactured wrong', 'wrong side',
                                  'thread', 'no thread']):
        return 'Defect: Broken / Structural Failure', 0.7

    if any(kw in text for kw in ['warranty', 'warranty claim']):
        return 'Defect: Malfunctions / Stops Working', 0.5

    if any(kw in text for kw in ['troubleshooting', 'troubleshoot']):
        return 'Defect: Malfunctions / Stops Working', 0.5

    if any(kw in text for kw in ['refund', 'return', 'exchange', 'replacement']):
        return 'Customer: Changed Mind / No Longer Needed', 0.3

    return 'Other / Miscellaneous', 0.0


def is_quality_category(category: str) -> bool:
    return category in QUALITY_CATEGORIES


def _conf_badge(conf: float) -> str:
    """Return an emoji confidence badge for table display."""
    if conf >= 0.85:
        return "🟢"
    elif conf >= 0.5:
        return "🟡"
    else:
        return "🔴"


# ─── AI Re-categorization ─────────────────────────────────────────────────────

def ai_recategorize_tickets(df: pd.DataFrame, analyzer, mode: str = 'uncertain') -> pd.DataFrame:
    """
    Use EnhancedAIAnalyzer to re-categorize tickets.
    mode='uncertain' → only tickets with Confidence < 0.5
    mode='all'       → every ticket
    Returns updated DataFrame.
    """
    if not _AI_IMPORT_OK or analyzer is None:
        return df

    mask = df['Confidence'] < 0.5 if mode == 'uncertain' else pd.Series(True, index=df.index)
    targets = df[mask]
    if targets.empty:
        return df

    updated = df.copy()
    progress = st.progress(0)
    n = len(targets)

    for i, (idx, row) in enumerate(targets.iterrows()):
        issue = str(row.get('Issue', '')) if pd.notna(row.get('Issue')) else ''
        if issue and len(issue) > 2:
            try:
                cat, conf, _, _ = analyzer.categorize_return(issue)
                updated.at[idx, 'Category'] = cat
                updated.at[idx, 'Confidence'] = conf
                updated.at[idx, 'Is Quality Issue'] = is_quality_category(cat)
            except Exception as e:
                logger.error(f"AI ticket re-cat error: {e}")
        progress.progress((i + 1) / n)

    progress.empty()
    return updated


# ─── Data Loading & Cleaning ─────────────────────────────────────────────────

def load_zendesk_data(uploaded_file) -> Optional[pd.DataFrame]:
    try:
        raw = pd.read_excel(uploaded_file)
        required = {"Ticket created - Date", "Ticket ID", "SKU", "Issue", "Ticket Type"}
        missing = required - set(raw.columns)
        if missing:
            st.error(f"Missing expected columns: {', '.join(missing)}")
            return None
        df = raw.drop_duplicates(subset="Ticket ID").copy()
        df["Ticket created - Date"] = pd.to_datetime(df["Ticket created - Date"], errors="coerce")
        return df
    except Exception as e:
        st.error(f"Failed to read file: {e}")
        return None


def filter_by_date(df: pd.DataFrame, start, end) -> pd.DataFrame:
    mask = (df["Ticket created - Date"] >= pd.Timestamp(start)) & (
        df["Ticket created - Date"] <= pd.Timestamp(end) + timedelta(hours=23, minutes=59, seconds=59)
    )
    return df.loc[mask].copy()


# ─── Categorize All Tickets ───────────────────────────────────────────────────

def categorize_all_tickets(df: pd.DataFrame) -> pd.DataFrame:
    categories, confidences = [], []
    for _, row in df.iterrows():
        issue = str(row.get("Issue", "")) if pd.notna(row.get("Issue")) else ""
        cat, conf = categorize_by_keywords(issue)
        categories.append(cat)
        confidences.append(conf)
    df = df.copy()
    df["Category"] = categories
    df["Confidence"] = confidences
    df["Is Quality Issue"] = df["Category"].apply(is_quality_category)
    df["Conf Badge"] = df["Confidence"].apply(_conf_badge)
    # Injury flag per ticket
    df["Safety Flag"] = df["Category"] == "Medical / Safety Concern"
    return df


# ─── Trend Analysis ───────────────────────────────────────────────────────────

def build_trend_data(df: pd.DataFrame, freq: str = 'W') -> pd.DataFrame:
    """Return weekly or monthly quality issue + safety concern counts."""
    qi = df[df["Is Quality Issue"] == True].copy()
    if qi.empty:
        return pd.DataFrame()
    qi["Period"] = qi["Ticket created - Date"].dt.to_period(freq).dt.start_time
    trend = qi.groupby("Period").size().reset_index(name="Quality Issues")
    safety = qi[qi["Category"] == "Medical / Safety Concern"].groupby("Period").size().reset_index(name="Safety Concerns")
    trend = trend.merge(safety, on="Period", how="left").fillna(0)
    trend["Safety Concerns"] = trend["Safety Concerns"].astype(int)
    return trend.set_index("Period")


# ─── Core Aggregation ─────────────────────────────────────────────────────────

def build_quality_report(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    qi = df[df["Is Quality Issue"] == True].copy()
    if qi.empty:
        return pd.DataFrame(), pd.DataFrame()

    qi = qi[qi["SKU"].notna()]
    qi = qi[~qi["SKU"].str.strip().isin(EXCLUDED_SKUS)]
    qi["Parent SKU"] = qi["SKU"].str[:PARENT_SKU_LENGTH]

    grouped = qi.groupby("Parent SKU", sort=False)

    agg = grouped.agg(
        Quality_Issue_Count=("Ticket ID", "count"),
        Unique_SKUs=("SKU", "nunique"),
        First_Seen=("Ticket created - Date", "min"),
        Last_Seen=("Ticket created - Date", "max"),
    ).reset_index()

    # Injury flag per parent SKU
    safety_skus = qi[qi["Category"] == "Medical / Safety Concern"]["Parent SKU"].unique()
    agg["Has Safety Concern"] = agg["Parent SKU"].isin(safety_skus)

    def sku_breakdown(grp):
        counts = grp["SKU"].value_counts()
        return "; ".join(f"{sku} x{cnt}" for sku, cnt in counts.items())

    sku_notes = grouped.apply(sku_breakdown, include_groups=False).reset_index()
    sku_notes.columns = ["Parent SKU", "SKU Breakdown"]

    def category_breakdown(grp):
        counts = grp["Category"].value_counts()
        return "; ".join(f"{cat} ({cnt})" for cat, cnt in counts.items())

    cat_notes = grouped.apply(category_breakdown, include_groups=False).reset_index()
    cat_notes.columns = ["Parent SKU", "Category Breakdown"]

    def top_issues(grp, n=3):
        counts = grp["Issue"].value_counts().head(n)
        return "\n".join(f"- {issue} ({cnt})" for issue, cnt in counts.items())

    issue_notes = grouped.apply(top_issues, include_groups=False).reset_index()
    issue_notes.columns = ["Parent SKU", "Top Issues"]

    def type_mix(grp):
        counts = grp["Ticket Type"].value_counts()
        return "; ".join(f"{t}: {c}" for t, c in counts.items())

    type_notes = grouped.apply(type_mix, include_groups=False).reset_index()
    type_notes.columns = ["Parent SKU", "Ticket Type Breakdown"]

    def source_mix(grp):
        if "Order source" not in grp.columns:
            return ""
        counts = grp["Order source"].value_counts()
        return "; ".join(f"{s}: {c}" for s, c in counts.items())

    source_notes = grouped.apply(source_mix, include_groups=False).reset_index()
    source_notes.columns = ["Parent SKU", "Order Source Breakdown"]

    rate_cols = {}
    if "Return completed?" in qi.columns:
        rate_cols["Returns_Completed"] = ("Return completed?", "sum")
    if "Replacement SO" in qi.columns:
        rate_cols["Replacements_Sent"] = ("Replacement SO", lambda x: (x.str.strip() != "").sum())

    if rate_cols:
        rate_agg = grouped.agg(**rate_cols).reset_index()
    else:
        rate_agg = agg[["Parent SKU"]].copy()
        rate_agg["Returns_Completed"] = 0
        rate_agg["Replacements_Sent"] = 0

    report = (
        agg
        .merge(sku_notes, on="Parent SKU")
        .merge(cat_notes, on="Parent SKU")
        .merge(issue_notes, on="Parent SKU")
        .merge(type_notes, on="Parent SKU")
        .merge(source_notes, on="Parent SKU")
        .merge(rate_agg, on="Parent SKU")
    )

    report = report.sort_values("Quality_Issue_Count", ascending=False).reset_index(drop=True)
    report.index = report.index + 1
    report.index.name = "Rank"

    report = report.rename(columns={
        "Quality_Issue_Count": "Quality Issues",
        "Unique_SKUs": "Variant Count",
        "First_Seen": "First Seen",
        "Last_Seen": "Last Seen",
        "Returns_Completed": "Returns Completed",
        "Replacements_Sent": "Replacements Sent",
    })

    # Add safety flag prefix to SKUs with safety concerns
    report["Parent SKU Display"] = report.apply(
        lambda r: f"🚨 {r['Parent SKU']}" if r["Has Safety Concern"] else r["Parent SKU"], axis=1
    )

    cat_summary = (
        qi.groupby("Category")
        .agg(Issue_Count=("Ticket ID", "count"), Products_Affected=("Parent SKU", "nunique"))
        .sort_values("Issue_Count", ascending=False)
        .reset_index()
    )
    cat_summary.index = cat_summary.index + 1
    cat_summary.index.name = "Rank"
    cat_summary = cat_summary.rename(columns={
        "Issue_Count": "Total Issues",
        "Products_Affected": "Products Affected",
    })
    cat_summary["Safety"] = cat_summary["Category"].apply(
        lambda c: "🚨" if c == "Medical / Safety Concern" else ""
    )

    return report, cat_summary


# ─── KPI helpers ──────────────────────────────────────────────────────────────

def compute_kpis(df: pd.DataFrame, report: pd.DataFrame) -> Dict:
    total_tickets = len(df)
    quality_tickets = int(df["Is Quality Issue"].sum())
    quality_rate = quality_tickets / total_tickets * 100 if total_tickets else 0
    products_affected = len(report)
    top_product = report.iloc[0]["Parent SKU"] if not report.empty else "N/A"
    top_count = int(report.iloc[0]["Quality Issues"]) if not report.empty else 0
    low_conf = int((df["Is Quality Issue"] & (df["Confidence"] < 0.5)).sum())
    injury_count = int(df["Safety Flag"].sum()) if "Safety Flag" in df.columns else 0
    return {
        "total_tickets": total_tickets,
        "quality_tickets": quality_tickets,
        "quality_rate": quality_rate,
        "products_affected": products_affected,
        "top_product": top_product,
        "top_count": top_count,
        "low_confidence_count": low_conf,
        "injury_count": injury_count,
    }


# ─── HTML Export (print-to-PDF ready) ────────────────────────────────────────

def export_html_report(report: pd.DataFrame, cat_summary: pd.DataFrame,
                       kpis: dict, date_label: str) -> str:
    """Generate a standalone HTML file suitable for browser print-to-PDF."""
    rows_html = ""
    for _, row in report.iterrows():
        injury = row.get("Has Safety Concern", False)
        row_style = ' style="background:#fff3cd;"' if injury else ""
        flag = "🚨 " if injury else ""
        rows_html += f"""<tr{row_style}>
          <td>{flag}{row['Parent SKU']}</td>
          <td style="text-align:center"><strong>{int(row['Quality Issues'])}</strong></td>
          <td style="font-size:0.8em">{row.get('Category Breakdown', '')}</td>
          <td style="font-size:0.8em">{row.get('Top Issues', '').replace(chr(10), '<br>')}</td>
          <td>{str(row.get('First Seen', ''))[:10]}</td>
          <td>{str(row.get('Last Seen', ''))[:10]}</td>
        </tr>"""

    cat_rows = ""
    for _, row in cat_summary.iterrows():
        safety = row.get("Safety", "")
        row_style = ' style="background:#fff3cd;"' if safety else ""
        cat_rows += f"""<tr{row_style}>
          <td>{safety} {row['Category']}</td>
          <td style="text-align:center"><strong>{int(row['Total Issues'])}</strong></td>
          <td style="text-align:center">{int(row['Products Affected'])}</td>
        </tr>"""

    injury_banner = ""
    if kpis.get("injury_count", 0) > 0:
        injury_banner = f"""<div style="background:#fff3cd;border:2px solid #ff6b35;
            border-radius:8px;padding:12px 16px;margin:12px 0;">
          <strong>🚨 SAFETY ALERT:</strong> {kpis['injury_count']} tickets classified as
          <em>Medical / Safety Concern</em>. Review immediately and open a Quality Issue in Odoo.
        </div>"""

    return f"""<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8">
<title>B2C Zendesk Quality Report — {date_label}</title>
<style>
  body{{font-family:Arial,sans-serif;color:#333;margin:0;padding:20px;font-size:13px}}
  h1{{background:#004366;color:#fff;padding:14px 20px;border-radius:8px;font-size:1.3em;margin-bottom:8px}}
  h2{{color:#004366;border-bottom:2px solid #23b2be;padding-bottom:4px;margin-top:24px}}
  .kpis{{display:flex;gap:12px;margin:12px 0;flex-wrap:wrap}}
  .kpi{{background:#f0fafb;border:1px solid #23b2be;border-radius:8px;padding:10px 16px;text-align:center;min-width:110px}}
  .kpi-val{{font-size:1.6em;font-weight:700;color:#004366}}
  .kpi-lbl{{font-size:0.75em;color:#666;margin-top:2px}}
  .kpi-inj{{border-color:#ff6b35;background:#fff3cd}}
  table{{width:100%;border-collapse:collapse;margin:10px 0}}
  th{{background:#23b2be;color:#fff;padding:7px 8px;text-align:left;font-size:0.85em}}
  td{{padding:5px 8px;border-bottom:1px solid #eee;vertical-align:top}}
  tr:nth-child(even){{background:#f9feff}}
  .footer{{margin-top:20px;font-size:0.75em;color:#aaa;text-align:center}}
  @media print{{body{{padding:0}}.no-print{{display:none}}}}
</style></head><body>
<h1>🎫 B2C Zendesk Quality Report — {date_label}</h1>
{injury_banner}
<div class="kpis">
  <div class="kpi"><div class="kpi-val">{kpis['total_tickets']:,}</div><div class="kpi-lbl">Total Tickets</div></div>
  <div class="kpi"><div class="kpi-val">{kpis['quality_tickets']:,}</div><div class="kpi-lbl">Quality Issues</div></div>
  <div class="kpi"><div class="kpi-val">{kpis['quality_rate']:.1f}%</div><div class="kpi-lbl">Quality Rate</div></div>
  <div class="kpi"><div class="kpi-val">{kpis['products_affected']}</div><div class="kpi-lbl">Products Affected</div></div>
  <div class="kpi{'kpi-inj' if kpis.get('injury_count',0)>0 else ''}"><div class="kpi-val">{kpis.get('injury_count',0)}</div><div class="kpi-lbl">Safety Concerns</div></div>
</div>
<p style="color:#666;font-size:0.8em">Generated {datetime.now().strftime('%Y-%m-%d %H:%M')} | Vive Health Quality Suite</p>
<h2>Quality Issues by Product (sorted by occurrence)</h2>
<table><thead><tr><th>Parent SKU</th><th>Issues</th><th>Categories</th><th>Top Issues</th><th>First Seen</th><th>Last Seen</th></tr></thead>
<tbody>{rows_html}</tbody></table>
<h2>Quality Issues by Category</h2>
<table><thead><tr><th>Category</th><th>Total Issues</th><th>Products Affected</th></tr></thead>
<tbody>{cat_rows}</tbody></table>
<div class="footer">Vive Health Quality Suite — Confidential</div>
</body></html>"""


# ─── Excel Export ─────────────────────────────────────────────────────────────

def export_report_xlsx(report: pd.DataFrame, cat_summary: pd.DataFrame,
                       kpis: dict, date_label: str) -> bytes:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    buf = BytesIO()
    export_report = report.drop(columns=["Has Safety Concern", "Parent SKU Display"], errors="ignore")
    export_cat = cat_summary.drop(columns=["Safety"], errors="ignore")

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        export_report.to_excel(writer, sheet_name="Quality by Product", startrow=4)
        wb = writer.book
        ws = writer.sheets["Quality by Product"]

        header_fill = PatternFill("solid", fgColor="004366")
        header_font = Font(name="Arial", bold=True, color="FFFFFF", size=14)
        ws.merge_cells("A1:L1")
        ws["A1"].value = f"B2C Zendesk Quality Report — {date_label}"
        ws["A1"].font = header_font
        ws["A1"].fill = header_fill
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 36

        sub_font = Font(name="Arial", size=10, color="004366")
        ws.merge_cells("A2:L2")
        ws["A2"].value = (
            f"Total Tickets: {kpis['total_tickets']:,}  |  "
            f"Quality Issues: {kpis['quality_tickets']:,}  ({kpis['quality_rate']:.1f}%)  |  "
            f"Products Affected: {kpis['products_affected']}  |  "
            f"Safety Concerns: {kpis.get('injury_count', 0)}  |  "
            f"Top Product: {kpis['top_product']} ({kpis['top_count']})"
        )
        ws["A2"].font = sub_font
        ws["A2"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A3:L3")
        ws["A3"].value = f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws["A3"].font = Font(name="Arial", size=9, italic=True, color="666666")
        ws["A3"].alignment = Alignment(horizontal="center")

        col_header_fill = PatternFill("solid", fgColor="23B2BE")
        col_header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        injury_fill = PatternFill("solid", fgColor="FFF3CD")
        thin_border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=5, column=col_idx)
            cell.font = col_header_font
            cell.fill = col_header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border

        # Map original parent SKU column to find safety rows
        safety_skus = set(report[report["Has Safety Concern"] == True]["Parent SKU"].tolist()) if "Has Safety Concern" in report.columns else set()

        alt_fill = PatternFill("solid", fgColor="F0FAFB")
        data_font = Font(name="Arial", size=10)
        for row_idx in range(6, ws.max_row + 1):
            # Check if this row's Parent SKU is a safety concern
            sku_cell = ws.cell(row=row_idx, column=2).value  # Parent SKU is column B (index 2)
            is_safety = str(sku_cell) in safety_skus if sku_cell else False
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if is_safety:
                    cell.fill = injury_fill
                elif row_idx % 2 == 0:
                    cell.fill = alt_fill

        widths = {"A": 6, "B": 12, "C": 13, "D": 11, "E": 13, "F": 13,
                  "G": 34, "H": 38, "I": 48, "J": 34, "K": 28, "L": 15, "M": 15}
        for col_letter, w in widths.items():
            ws.column_dimensions[col_letter].width = w
        ws.freeze_panes = "A6"

        export_cat.to_excel(writer, sheet_name="Quality by Category", startrow=2)
        ws2 = writer.sheets["Quality by Category"]
        ws2.merge_cells("A1:D1")
        ws2["A1"].value = "Quality Issues by Category"
        ws2["A1"].font = Font(name="Arial", bold=True, color="004366", size=13)
        ws2["A1"].alignment = Alignment(horizontal="center")
        for col_idx in range(1, 5):
            cell = ws2.cell(row=3, column=col_idx)
            cell.font = col_header_font
            cell.fill = col_header_fill
            cell.alignment = Alignment(horizontal="center")
        for col_letter, w in {"A": 6, "B": 42, "C": 14, "D": 18}.items():
            ws2.column_dimensions[col_letter].width = w

    return buf.getvalue()


# ─── Module Description ───────────────────────────────────────────────────────

MODULE_DESCRIPTION = """
<div style="background: rgba(0, 217, 255, 0.08); border: 1px solid #23b2be;
            border-radius: 10px; padding: 1rem 1.2rem; margin-bottom: 1.2rem;">
    <strong style="color:#004366;">📌 Purpose:</strong>
    Categorize <strong>all</strong> Zendesk tickets by Issue text using the same
    medical-device quality categories as the Return Categorizer, then report
    only quality-relevant issues grouped by <strong>Parent SKU</strong> (first 7 chars).<br>
    <strong style="color:#004366;">📊 Output:</strong>
    Tables sorted by issue count with category breakdowns, trend charts,
    cross-reference with returns data, and print-ready HTML export.
    Does <strong>not</strong> rely on the <em>Quality Issues?</em> checkbox.
</div>
"""

VIVE_TEAL = "#23b2be"
VIVE_NAVY = "#004366"


# ─── Streamlit UI ─────────────────────────────────────────────────────────────

def render_b2b_zendesk_reporting():
    """Main render function."""

    st.markdown("### 🎫 B2C Zendesk Quality Reporting")
    st.markdown(MODULE_DESCRIPTION, unsafe_allow_html=True)

    # ── File upload ────────────────────────────────────────────────────────
    uploaded = st.file_uploader(
        "Upload B2C Quality Issues Recordings (.xlsx)",
        type=["xlsx", "xls"],
        key="zendesk_uploader",
        help="Upload the B2C_QUALITY_ISSUES_RECORDINGS export from Zendesk.",
    )

    if not uploaded:
        st.info("Upload the **B2C_QUALITY_ISSUES_RECORDINGS** file to get started.")
        return

    with st.spinner("Reading and deduplicating Zendesk data..."):
        df = load_zendesk_data(uploaded)

    if df is None or df.empty:
        return

    st.success(f"Loaded **{len(df):,}** unique tickets")

    # ── Date range ────────────────────────────────────────────────────────
    st.markdown("#### 📅 Date Range")
    date_col = "Ticket created - Date"
    min_date = df[date_col].min().date()
    max_date = df[date_col].max().date()

    col_opt, col_range = st.columns([1, 3])
    with col_opt:
        use_all = st.checkbox("Use all dates", value=True, key="zendesk_all_dates")
    with col_range:
        if use_all:
            start_date, end_date = min_date, max_date
            st.caption(f"Full range: **{min_date}** to **{max_date}**")
        else:
            start_date, end_date = st.date_input(
                "Select range", value=(min_date, max_date),
                min_value=min_date, max_value=max_date,
                key="zendesk_date_range",
            )

    filtered = filter_by_date(df, start_date, end_date)
    st.info(f"**{len(filtered):,}** tickets in selected range")

    # ── Categorization mode ───────────────────────────────────────────────
    st.markdown("#### ⚙️ Categorization Mode")
    cat_mode = st.radio(
        "How should tickets be categorized?",
        options=[
            "🔑 Keyword Matching (instant, free)",
            "🔑+🤖 Keywords + AI for unclear tickets (recommended)",
            "🤖 Full AI Analysis (all tickets via API)",
        ],
        index=0, key="zendesk_cat_mode", horizontal=True,
    )

    # ── Generate ──────────────────────────────────────────────────────────
    if st.button("🚀 Generate Quality Report", type="primary", key="zendesk_run"):
        for k in ["zendesk_report", "zendesk_cat_summary", "zendesk_kpis",
                  "zendesk_date_label", "zendesk_categorized"]:
            st.session_state.pop(k, None)

        with st.spinner("Categorizing all tickets with keyword matching..."):
            categorized = categorize_all_tickets(filtered)

        qi_count = int(categorized["Is Quality Issue"].sum())
        st.toast(f"Keyword pass: {qi_count} quality issues identified")

        # ── AI upgrade passes ──────────────────────────────────────────────
        if "🤖" in cat_mode:
            if _AI_IMPORT_OK:
                try:
                    # Try to get the AI analyzer from Streamlit session state
                    analyzer = st.session_state.get("ai_analyzer")
                    if analyzer is None:
                        from enhanced_ai_analysis import EnhancedAIAnalyzer, AIProvider
                        analyzer = EnhancedAIAnalyzer(AIProvider.FASTEST)
                    mode = 'all' if "Full AI" in cat_mode else 'uncertain'
                    target_n = len(categorized) if mode == 'all' else int((categorized["Confidence"] < 0.5).sum())
                    with st.spinner(f"AI recategorizing {target_n} tickets..."):
                        categorized = ai_recategorize_tickets(categorized, analyzer, mode=mode)
                    st.toast(f"AI pass complete on {target_n} tickets")
                except Exception as e:
                    st.warning(f"AI unavailable ({e}). Using keyword results.")
            else:
                st.warning("AI modules not loaded. Showing keyword-only results.")

        with st.spinner("Building quality report..."):
            report, cat_summary = build_quality_report(categorized)
            kpis = compute_kpis(categorized, report)
            date_label = (
                f"{start_date.strftime('%b %d')} - {end_date.strftime('%b %d, %Y')}"
                if not use_all else f"{min_date.strftime('%b %Y')} (All Data)"
            )
            st.session_state["zendesk_report"] = report
            st.session_state["zendesk_cat_summary"] = cat_summary
            st.session_state["zendesk_kpis"] = kpis
            st.session_state["zendesk_date_label"] = date_label
            st.session_state["zendesk_categorized"] = categorized

    # ── Display ───────────────────────────────────────────────────────────
    if st.session_state.get("zendesk_report") is not None:
        report = st.session_state["zendesk_report"]
        cat_summary = st.session_state["zendesk_cat_summary"]
        kpis = st.session_state["zendesk_kpis"]
        date_label = st.session_state["zendesk_date_label"]
        categorized = st.session_state["zendesk_categorized"]

        if report.empty:
            st.warning("No quality issues found for the selected parameters.")
            return

        # ── Injury alert banner ────────────────────────────────────────────
        if kpis.get("injury_count", 0) > 0:
            st.error(
                f"🚨 **SAFETY ALERT: {kpis['injury_count']} ticket(s) classified as Medical / Safety Concern.** "
                "These require immediate review and a Quality Issue opened in Odoo. "
                "Safety rows are highlighted in yellow throughout this report."
            )

        # ── KPI cards ─────────────────────────────────────────────────────
        st.markdown("---")
        st.markdown(f"#### 📊 Quality Summary — {date_label}")
        k1, k2, k3, k4, k5, k6 = st.columns(6)
        with k1: st.metric("Total Tickets", f"{kpis['total_tickets']:,}")
        with k2: st.metric("Quality Issues", f"{kpis['quality_tickets']:,}")
        with k3: st.metric("Quality Rate", f"{kpis['quality_rate']:.1f}%")
        with k4: st.metric("Products Affected", kpis["products_affected"])
        with k5: st.metric("Low Confidence", kpis["low_confidence_count"],
                           help="Tickets with confidence < 50%")
        with k6: st.metric("Safety Concerns", kpis.get("injury_count", 0),
                           help="Medical / Safety Concern tickets")

        # ── Safety concern detail table ───────────────────────────────────
        if kpis.get("injury_count", 0) > 0:
            st.markdown("---")
            st.markdown(
                "<div style='background:#fff0f0;border:2px solid #cc0000;"
                "border-radius:8px;padding:0.8rem 1rem;margin-bottom:0.5rem'>"
                "<span style='color:#cc0000;font-size:1.1rem;font-weight:700'>"
                f"🚨 {kpis['injury_count']} Medical / Safety Concern Ticket(s) — Requires Immediate Action"
                "</span></div>",
                unsafe_allow_html=True,
            )
            safety_tickets = categorized[
                categorized["Category"] == "Medical / Safety Concern"
            ].copy()
            safety_cols = [c for c in [
                "Ticket ID", "Ticket created - Date", "Parent SKU", "SKU",
                "Issue", "Category", "Confidence",
                "Ticket Type", "Order Source",
            ] if c in safety_tickets.columns]
            st.dataframe(
                safety_tickets[safety_cols].sort_values(
                    "Ticket created - Date", ascending=False
                ) if "Ticket created - Date" in safety_cols else safety_tickets[safety_cols],
                width="stretch",
                height=min(400, 38 + 35 * len(safety_tickets)),
                column_config={
                    "Ticket created - Date": st.column_config.DateColumn("Date", format="YYYY-MM-DD"),
                    "Confidence": st.column_config.NumberColumn("Confidence", format="%.0%%"),
                    "Issue": st.column_config.TextColumn("Issue", width="large"),
                },
            )
            st.caption("⬆ These tickets must be reviewed and a Quality Issue opened in Odoo.")

        # ── Top products by total issues ──────────────────────────────────
        st.markdown("---")
        st.markdown("#### 🏆 Top Products by Quality Issue Volume")
        top_n = min(15, len(report))
        top_sorted = (
            report.sort_values("Quality Issues", ascending=True)
            .tail(top_n)
            .set_index("Parent SKU Display")[["Quality Issues"]]
        )
        st.bar_chart(top_sorted, color=VIVE_TEAL, horizontal=True)
        st.caption("Sorted by total quality issues — longest bar = most problems.")

        # ── Category summary ──────────────────────────────────────────────
        st.markdown("---")
        st.markdown("#### 📂 Quality Issues by Category — All Products")
        col_table, col_chart = st.columns([1, 1])
        with col_table:
            st.dataframe(
                cat_summary.sort_values("Total Issues", ascending=False),
                width='stretch',
                height=min(500, 38 + 35 * len(cat_summary)),
                column_config={
                    "Total Issues": st.column_config.NumberColumn(format="%d"),
                    "Products Affected": st.column_config.NumberColumn(format="%d"),
                    "Safety": st.column_config.TextColumn("⚠️", width="small"),
                },
            )
        with col_chart:
            cat_chart_df = (
                cat_summary.sort_values("Total Issues", ascending=True)
                .set_index("Category")[["Total Issues"]]
            )
            st.bar_chart(cat_chart_df, color=VIVE_TEAL, horizontal=True)

        # ── Product table with filters ────────────────────────────────────
        st.markdown("---")
        st.markdown("#### 📋 Quality Issues by Product")

        # Filters
        with st.expander("🔧 Filter & Sort Options", expanded=False):
            fc1, fc2, fc3 = st.columns(3)
            with fc1:
                sku_search = st.text_input("SKU contains", key="zendesk_sku_filter",
                                           placeholder="e.g. MOB1027")
            with fc2:
                all_cats_in_report = sorted(set(
                    cat.strip()
                    for breakdown in report["Category Breakdown"].dropna()
                    for cat in breakdown.split(";")
                    if cat.strip()
                ))
                # Extract just the category name (before the count in parentheses)
                clean_cats = sorted(set(
                    re.sub(r'\s*\(\d+\)\s*$', '', c).strip()
                    for c in all_cats_in_report if c
                ))
                cat_filter = st.multiselect("Category contains", clean_cats,
                                            key="zendesk_cat_filter")
            with fc3:
                min_issues = st.slider("Min quality issues", 1,
                                       max(1, int(report["Quality Issues"].max())),
                                       1, key="zendesk_min_issues")

        display_report = report.copy()
        if sku_search:
            display_report = display_report[
                display_report["Parent SKU"].str.contains(sku_search, case=False, na=False)
            ]
        if cat_filter:
            def has_any_cat(breakdown):
                if pd.isna(breakdown):
                    return False
                return any(c in breakdown for c in cat_filter)
            display_report = display_report[display_report["Category Breakdown"].apply(has_any_cat)]
        display_report = display_report[display_report["Quality Issues"] >= min_issues]

        display_cols = [
            "Parent SKU Display", "Quality Issues", "Variant Count",
            "Category Breakdown", "SKU Breakdown", "Top Issues",
            "Ticket Type Breakdown", "Order Source Breakdown",
            "Returns Completed", "Replacements Sent",
            "First Seen", "Last Seen",
        ]
        available_cols = [c for c in display_cols if c in display_report.columns]

        st.dataframe(
            display_report[available_cols].sort_values("Quality Issues", ascending=False),
            width='stretch',
            height=min(600, 38 + 35 * len(display_report)),
            column_config={
                "Parent SKU Display": st.column_config.TextColumn("Parent SKU"),
                "Quality Issues": st.column_config.NumberColumn(format="%d"),
                "Variant Count": st.column_config.NumberColumn("Variants", format="%d"),
                "First Seen": st.column_config.DateColumn(format="YYYY-MM-DD"),
                "Last Seen": st.column_config.DateColumn(format="YYYY-MM-DD"),
                "Returns Completed": st.column_config.NumberColumn(format="%d"),
                "Replacements Sent": st.column_config.NumberColumn(format="%d"),
                "Category Breakdown": st.column_config.TextColumn("Category Breakdown", width="large"),
                "SKU Breakdown": st.column_config.TextColumn("SKU Breakdown", width="large"),
                "Top Issues": st.column_config.TextColumn("Top Issues", width="large"),
                "Ticket Type Breakdown": st.column_config.TextColumn("Ticket Types", width="medium"),
                "Order Source Breakdown": st.column_config.TextColumn("Order Sources", width="medium"),
            },
        )
        st.caption(
            f"Showing **{len(display_report)}** of **{len(report)}** parent SKUs. "
            "Click any column header to sort. 🚨 = Safety Concern flagged."
        )

        if len(display_report) > 1:
            st.markdown("#### 🏷️ Filtered View — Top Products by Issue Count")
            top_n_filtered = min(15, len(display_report))
            top_chart_filtered = (
                display_report
                .sort_values("Quality Issues", ascending=True)
                .tail(top_n_filtered)
                .set_index("Parent SKU Display")[["Quality Issues"]]
            )
            st.bar_chart(top_chart_filtered, color=VIVE_TEAL, horizontal=True)
            st.caption("Reflects active filters above.")

        # ── Cross-reference with Return Categorizer ───────────────────────
        with st.expander("🔗 Cross-Reference with Return Categorizer Data", expanded=False):
            returns_data = st.session_state.get("categorized_data")
            col_mapping = st.session_state.get("column_mapping", {})
            if returns_data is not None and col_mapping:
                sku_col = col_mapping.get("sku")
                cat_col = col_mapping.get("category")
                if sku_col and cat_col and sku_col in returns_data.columns:
                    st.markdown("Comparing Zendesk ticket categories vs. Amazon return categories for matching Parent SKUs.")
                    # Build returns summary by parent SKU
                    ret = returns_data.copy()
                    ret["Parent SKU"] = ret[sku_col].astype(str).str[:PARENT_SKU_LENGTH]
                    ret_summary = (
                        ret[ret[cat_col].notna() & (ret[cat_col] != "")]
                        .groupby("Parent SKU")[cat_col]
                        .agg(lambda x: x.dropna().value_counts().index[0] if len(x.dropna()) > 0 else "")
                        .reset_index()
                        .rename(columns={cat_col: "Top Return Category"})
                    )
                    # Zendesk top category per SKU
                    zen_top = (
                        categorized[categorized["Is Quality Issue"] == True].copy()
                    )
                    zen_top["Parent SKU"] = zen_top["SKU"].str[:PARENT_SKU_LENGTH]
                    zen_summary = (
                        zen_top.groupby("Parent SKU")["Category"]
                        .agg(lambda x: x.dropna().value_counts().index[0] if len(x.dropna()) > 0 else "")
                        .reset_index()
                        .rename(columns={"Category": "Top Zendesk Category"})
                    )
                    cross = zen_summary.merge(ret_summary, on="Parent SKU", how="inner")
                    cross["Match?"] = cross.apply(
                        lambda r: "✅ Match" if r["Top Zendesk Category"] == r["Top Return Category"] else "🔀 Differs",
                        axis=1
                    )
                    cross = cross.sort_values("Match?")
                    st.dataframe(cross, width='stretch')
                    matches = (cross["Match?"] == "✅ Match").sum()
                    st.caption(
                        f"{matches}/{len(cross)} SKUs have matching top categories between Zendesk and Returns data."
                    )
                else:
                    st.info("Return data is loaded but SKU/Category columns could not be mapped.")
            else:
                st.info(
                    "Run the **Return Categorizer** tool first to load returns data, "
                    "then return here to see the cross-reference comparison."
                )

        # ── Drill-down expanders ───────────────────────────────────────────
        with st.expander("🔎 Drill Down — All Categorized Quality Tickets", expanded=False):
            qi_detail = categorized[categorized["Is Quality Issue"] == True].copy()
            if "SKU" in qi_detail.columns:
                qi_detail["Parent SKU"] = qi_detail["SKU"].str[:PARENT_SKU_LENGTH]
            detail_cols = [
                "Ticket created - Date", "Ticket ID", "Parent SKU", "SKU",
                "Conf Badge", "Category", "Confidence", "Issue", "Ticket Type",
            ]
            avail = [c for c in detail_cols if c in qi_detail.columns]
            # Sort BEFORE subsetting columns so sort keys don't need to be in avail
            _sort_cols = [c for c in ["Safety Flag", "Ticket created - Date"] if c in qi_detail.columns]
            if _sort_cols:
                qi_detail_sorted = qi_detail.sort_values(
                    _sort_cols, ascending=[False] * len(_sort_cols)
                )[avail]
            else:
                qi_detail_sorted = qi_detail[avail]
            st.dataframe(qi_detail_sorted, width='stretch', height=400,
                         column_config={
                             "Conf Badge": st.column_config.TextColumn("Conf", width="small"),
                             "Confidence": st.column_config.NumberColumn(format="%.2f"),
                         })

        with st.expander("🔍 Low Confidence Tickets — AI Review Candidates", expanded=False):
            low_conf_df = categorized[categorized["Confidence"] < 0.5].copy()
            if not low_conf_df.empty:
                lc_cols = ["Ticket ID", "Issue", "Conf Badge", "Category", "Confidence", "Ticket Type", "SKU"]
                avail_lc = [c for c in lc_cols if c in low_conf_df.columns]
                st.dataframe(low_conf_df[avail_lc].sort_values("Confidence"),
                             width='stretch', height=300,
                             column_config={
                                 "Conf Badge": st.column_config.TextColumn("Conf", width="small"),
                                 "Confidence": st.column_config.NumberColumn(format="%.2f"),
                             })
                st.caption(f"{len(low_conf_df)} tickets with confidence < 50% — consider running AI mode.")
            else:
                st.success("All tickets categorized with high confidence.")

        with st.expander("📊 Non-Quality Tickets (excluded from report)", expanded=False):
            non_qi = categorized[categorized["Is Quality Issue"] == False].copy()
            non_summary = non_qi["Category"].value_counts().reset_index()
            non_summary.columns = ["Category", "Count"]
            st.dataframe(non_summary, width='stretch')
            st.caption(f"{len(non_qi)} tickets classified as non-quality")

        # ── Export ────────────────────────────────────────────────────────
        st.markdown("---")
        col_dl1, col_dl2, col_dl3, col_dl4, col_clear = st.columns([2, 2, 2, 2, 1])

        with col_dl1:
            xlsx_bytes = export_report_xlsx(report, cat_summary, kpis, date_label)
            st.download_button(
                "⬇️ Download Report (.xlsx)",
                data=xlsx_bytes,
                file_name=f"B2C_Zendesk_Quality_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary", key="zendesk_dl_xlsx",
            )

        with col_dl2:
            html_str = export_html_report(report, cat_summary, kpis, date_label)
            st.download_button(
                "⬇️ Download HTML (Print to PDF)",
                data=html_str.encode("utf-8"),
                file_name=f"B2C_Zendesk_Quality_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.html",
                mime="text/html", key="zendesk_dl_html",
                help="Open in browser and use Ctrl+P / Cmd+P to save as PDF",
            )

        with col_dl3:
            csv_bytes = report.drop(
                columns=["Has Safety Concern", "Parent SKU Display"], errors="ignore"
            ).to_csv(index=True).encode("utf-8")
            st.download_button(
                "⬇️ Product Report (.csv)",
                data=csv_bytes,
                file_name=f"B2C_Zendesk_Product_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                mime="text/csv", key="zendesk_dl_csv",
            )

        with col_dl4:
            full_csv = categorized.drop(
                columns=["Conf Badge", "Safety Flag"], errors="ignore"
            ).to_csv(index=False).encode("utf-8")
            st.download_button(
                "⬇️ Full Categorized Data (.csv)",
                data=full_csv,
                file_name=f"B2C_Zendesk_All_Categorized_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                mime="text/csv", key="zendesk_dl_full",
            )

        with col_clear:
            if st.button("🔄 Reset", key="zendesk_clear"):
                for k in ["zendesk_report", "zendesk_cat_summary", "zendesk_kpis",
                          "zendesk_date_label", "zendesk_categorized"]:
                    st.session_state.pop(k, None)
                st.rerun()
